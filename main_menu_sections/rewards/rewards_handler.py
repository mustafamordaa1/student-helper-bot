from datetime import datetime
import os
from telegram import (
    InputMediaPhoto,
    InputMediaVideo,
    Update,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from telegram.ext import CallbackContext
import openpyxl
import logging

from config import REWARDS_DAILY_GIFTS, REWARDS_EXCEL
from utils.database import execute_query, get_data
from utils.subscription_management import check_subscription

# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)


async def handle_rewards(update: Update, context: CallbackContext):
    """Handles the 'المكافآت' option and displays its sub-menu."""
    if not await check_subscription(update, context):
        return

    context.user_data["current_section"] = "rewards"  # Set user context
    keyboard = [
        [InlineKeyboardButton("هديتك اليومية 🎁", callback_data="handle_daily_reward")],
        [
            InlineKeyboardButton(
                "مكافآت المذاكرة الفخمة 🏆", callback_data="handle_premium_rewards"
            )
        ],
        [InlineKeyboardButton("الرجوع للخلف 🔙", callback_data="go_back")],
    ]
    await update.callback_query.edit_message_text(
        "المكافآت ✨", reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def handle_daily_reward(update: Update, context: CallbackContext):
    """Handles the 'هديتك اليومية' sub-option."""
    await update.callback_query.answer()  # Acknowledge callback interaction

    # user_id = update.effective_user.id
    # if await has_user_claimed_daily_reward(user_id):
    #     await update.callback_query.message.reply_text(
    #         "لقد استلمت هديتك اليومية بالفعل! 🎉"
    #     )
    #     return

    await update.callback_query.message.reply_text("جاري إعداد هديتك اليومية... 🎁")

    # Get the current day of the month (1-31)
    current_day = datetime.now().day

    # Construct the folder path for the daily reward
    folder_path = os.path.join(REWARDS_DAILY_GIFTS, f"day_{current_day}/")

    # Create the 'daily_gifts' folder structure if it doesn't exist (only once)
    daily_gifts_path = REWARDS_DAILY_GIFTS
    if not os.path.exists(daily_gifts_path):
        create_daily_gifts_folders(daily_gifts_path)

    # Prepare a list to store media (images and videos)
    media_group = []
    message_text = ""

    # Iterate through files in the folder
    try:
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)

            if filename.endswith((".jpg", ".jpeg", ".png")):
                media_group.append(InputMediaPhoto(open(file_path, "rb")))
            elif filename.endswith((".mp4", ".mov")):
                media_group.append(InputMediaVideo(open(file_path, "rb")))
            elif filename.endswith(".txt"):
                with open(file_path, "r", encoding="utf-8") as f:
                    message_text = f.read()
    except FileNotFoundError:
        logger.error(f"Daily gift folder not found for day {current_day}")
        await update.callback_query.message.reply_text(
            "عذرًا، حدث خطأ أثناء جلب هديتك اليومية. حاول مرة أخرى لاحقًا. 😥"
        )
        return

    # Send the content to the user
    if media_group:
        await context.bot.send_media_group(
            chat_id=update.effective_chat.id, media=media_group
        )
    if message_text:
        await context.bot.send_message(
            chat_id=update.effective_chat.id, text=message_text
        )

    # Increment the user's daily gifts used count
    await increment_user_daily_gifts_used(update.effective_user.id)

    await update.callback_query.message.reply_text("تم عرض هديتك اليومية! 🎁")


def create_daily_gifts_folders(daily_gifts_path):
    """Creates the daily gifts folder structure (days 1-31)."""
    os.makedirs(daily_gifts_path)
    for day in range(1, 32):
        day_folder_path = os.path.join(daily_gifts_path, f"day_{day}")
        os.makedirs(day_folder_path)
        # Create a default message file for each day
        with open(
            os.path.join(day_folder_path, "message.txt"), "w", encoding="utf-8"
        ) as f:
            f.write(f"هذه هي هدية اليوم {day}! 🎁")


async def has_user_claimed_daily_reward(user_id):
    """Checks if the user has already claimed the daily reward today."""
    user_data = get_data(
        "SELECT last_daily_reward_claim FROM users WHERE telegram_id = ?", (user_id,)
    )

    if not user_data or not user_data[0][0]:  # No claim date or first time claim
        return False

    last_claim_date = datetime.strptime(user_data[0][0], "%Y-%m-%d").date()
    today = datetime.now().date()

    return last_claim_date == today


async def increment_user_daily_gifts_used(user_id):
    """Increments the user's number_of_daily_gifts_used in the database."""
    is_today = await has_user_claimed_daily_reward(user_id)
    if is_today:
        return
    execute_query(
        "UPDATE users SET number_of_daily_gifts_used = number_of_daily_gifts_used + 1, last_daily_reward_claim = ? WHERE telegram_id = ?",
        (datetime.now().strftime("%Y-%m-%d"), user_id),
    )


async def handle_premium_rewards(update: Update, context: CallbackContext):
    """Handles the 'مكافآت المذاكرة الفخمة' sub-option."""
    await update.callback_query.answer()  # Acknowledge callback interaction
    await update.callback_query.message.reply_text(
        "🕒 جاري جلب مكافآت المذاكرة الفخمة... "
    )

    user_id = update.effective_user.id
    user_stats = await get_user_stats(user_id)

    if not user_stats:
        await update.callback_query.message.reply_text(
            "⚠️ لا توجد إحصائيات متاحة. ابدأ الدراسة لكسب المكافآت!"
        )
        return

    reward_messages = await generate_reward_messages(user_stats)
    stats_text = format_stats_text(user_stats, reward_messages)

    await update.callback_query.message.reply_text(stats_text)


async def get_user_stats(user_id):
    """Fetches and formats user statistics from the database."""
    try:
        user_data = get_data(
            "SELECT points, percentage_expected, usage_time, total_number_of_created_questions FROM users WHERE telegram_id = ?",
            (user_id,),
        )
    except Exception as e:
        logger.error(f"Error fetching user stats: {e}")
        return None

    if not user_data:
        return None

    points, percentage_expected, usage_time_str, questions_created = user_data[0]
    usage_time = convert_usage_time_to_hours(usage_time_str)

    return {
        "points": points if points is not None else 0,
        "percentage": (
            round(percentage_expected) if percentage_expected is not None else 0
        ),
        "time_spent": usage_time,
        "questions_created": questions_created if questions_created is not None else 0,
    }


def convert_usage_time_to_hours(usage_time_str):
    """Converts usage time string (HH:MM:SS) to hours (float)."""
    if usage_time_str:
        try:
            hours, minutes, seconds = map(int, usage_time_str.split(":"))
            return round(hours + minutes / 60 + seconds / 3600, 2)
        except ValueError:
            logger.warning(f"Invalid usage time format: {usage_time_str}")
            return 0
    return 0


async def generate_reward_messages(user_stats):
    """Generates reward messages based on user statistics."""
    reward_messages = []

    try:
        workbook = openpyxl.load_workbook(REWARDS_EXCEL)
        rewards_sheet = workbook.active

        # Get target and reward text from the SECOND row (index 1)
        percentage_target = rewards_sheet.cell(row=2, column=1).value
        percentage_reward = rewards_sheet.cell(row=2, column=2).value
        time_target = rewards_sheet.cell(row=2, column=3).value
        time_reward = rewards_sheet.cell(row=2, column=4).value
        questions_target = rewards_sheet.cell(row=2, column=5).value
        questions_reward = rewards_sheet.cell(row=2, column=6).value
        points_target = rewards_sheet.cell(row=2, column=7).value
        points_reward = rewards_sheet.cell(row=2, column=8).value

        # Format the messages
        reward_messages.extend(
            [
                format_reward_message(
                    user_stats["percentage"],
                    percentage_target,
                    percentage_reward,
                    "النسبة المئوية",
                    "%",
                ),
                format_reward_message(
                    user_stats["time_spent"],
                    time_target,
                    time_reward,
                    "وقت الدراسة",
                    "ساعة",
                ),
                format_reward_message(
                    user_stats["questions_created"],
                    questions_target,
                    questions_reward,
                    "سؤال",
                    "سؤال",
                ),
                format_reward_message(
                    user_stats["points"], points_target, points_reward, "نقطة", "نقطة"
                ),
            ]
        )
    except FileNotFoundError:
        logger.error("rewards.xlsx file not found.")
    except Exception as e:
        logger.error(f"Error generating reward messages: {e}")

    return reward_messages


def format_reward_message(user_value, target_value, reward_text, metric_name, unit):
    """Formats a single reward message."""
    if user_value >= target_value:
        return f"✅ - {reward_text}: حققت {user_value} {unit} من {target_value} {unit} المطلوبة من {metric_name}.\n"
    else:
        return f"⚠️ - {reward_text}: تحتاج إلى {target_value - user_value} {unit} إضافية من {metric_name} للوصول إلى {target_value} {unit}.\n"


def format_stats_text(user_stats, reward_messages: list):
    """Formats the statistics and reward messages into a single string."""
    reward_messages = "".join(reward_messages)

    return (
        f"📊 إحصائياتك:\n"
        f"🏅 نقاطك: {user_stats['points']}\n"
        f"📈 النسبة المئوية: {user_stats['percentage']}%\n"
        f"⏳ وقت الدراسة: {user_stats['time_spent']} ساعة\n"
        f"✏️ عدد الأسئلة التي قمت بإنشائها: {user_stats['questions_created']}\n\n"
        f"🎁 مكافآتك:\n{reward_messages}"
    )


def format_no_rewards_text(user_stats):
    """Formats the message when the user has no rewards."""
    return (
        f"📊 إحصائياتك:\n"
        f"🏅 نقاطك: {user_stats['points']}\n"
        f"📈 النسبة المئوية: {user_stats['percentage']}%\n"
        f"⏳ وقت الدراسة: {user_stats['time_spent']} ساعة\n"
        f"✏️ عدد الأسئلة التي قمت بإنشائها: {user_stats['questions_created']}\n\n"
        f"🙅‍♂️ لم تكسب أي مكافآت حتى الآن. استمر في الدراسة!"
    )


# Dictionary to map handler names to functions
REWARDS_HANDLERS = {
    "rewards": handle_rewards,
    "handle_daily_reward": handle_daily_reward,
    "handle_premium_rewards": handle_premium_rewards,
}
