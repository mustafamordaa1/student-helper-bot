"""
Handles loading and reading data from Excel files.

This class provides methods for interacting with Excel workbooks, such as:
- Loading a workbook from a file path.
- Getting a list of sheet names.
- Retrieving data from a specific sheet.
- Reading the value of a specific cell.
"""

import openpyxl


class ExcelHandler:
    """Handles loading and reading data from Excel files."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = self._load_workbook()

    def _load_workbook(self):
        """Loads the workbook from the file path."""
        try:
            return openpyxl.load_workbook(self.file_path)
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return None

    def get_sheet_names(self) -> list:
        """Returns a list of sheet names from the workbook."""
        return self.workbook.sheetnames if self.workbook else []

    def get_sheet_data(self, sheet_name: str, min_row=2, values_only=True) -> list:
        """Returns data from a specific sheet."""
        if self.workbook:
            try:
                sheet = self.workbook[sheet_name]
                return list(sheet.iter_rows(min_row=min_row, values_only=values_only))
            except Exception as e:
                print(f"Error reading sheet: {e}")
                return []
        return []

    def get_cell_value(self, sheet_name: str, row: int, column: int):
        """Returns the value of a specific cell."""
        if self.workbook:
            try:
                sheet = self.workbook[sheet_name]
                return sheet.cell(row=row, column=column).value
            except Exception as e:
                print(f"Error accessing cell: {e}")
                return None
        return None
