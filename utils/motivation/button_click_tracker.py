import os
import random
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import CallbackContext
from config import (
    FEMALE_GO_BACK_MESSAGES_FILE,
    FEMALE_MAIN_MENUMESSAGES_FILE,
    MALE_GO_BACK_MESSAGES_FILE,
    MALE_MAIN_MENU_MESSAGES_FILE,
)
from utils import user_management
from utils.database import get_data

# Constants for paths and message frequency
MESSAGES_TRIGGER_THRESHOLD = 3

# Global dictionary to store messages (loaded on bot startup)
motivational_messages = {
    "main_menu": {"male": [], "female": []},
    "go_back": {"male": [], "female": []},
}


def load_motivational_messages():
    """Loads motivational messages from Excel files into the global dictionary."""

    def load_messages_from_file(filepath):
        messages = []
        workbook = load_workbook(filepath)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):  # Header row
            messages.append(row[0])  # Messages are in the first column
        return messages

    motivational_messages["main_menu"]["male"] = load_messages_from_file(
        MALE_MAIN_MENU_MESSAGES_FILE
    )
    motivational_messages["main_menu"]["female"] = load_messages_from_file(
        FEMALE_MAIN_MENUMESSAGES_FILE
    )

    motivational_messages["go_back"]["male"] = load_messages_from_file(
        MALE_GO_BACK_MESSAGES_FILE
    )
    motivational_messages["go_back"]["female"] = load_messages_from_file(
        FEMALE_GO_BACK_MESSAGES_FILE
    )


def get_random_motivational_message(gender: str, called_from):
    """Returns a random motivational message based on gender."""
    if gender.lower() == "male":
        return random.choice(motivational_messages[called_from]["male"])
    elif gender.lower() == "female":
        return random.choice(motivational_messages[called_from]["female"])
    else:
        return None


async def send_motivational_message(
    update: Update, context: CallbackContext, called_from
):
    """Sends a motivational message to the user and resets the click counter."""
    # user_id = update.effective_user.id

    if isinstance(update, Update):
        user = update.effective_user
    else:  # Assuming it's a CallbackQuery
        user = update.from_user
    user_id = user.id
    gender = get_data("SELECT gender FROM users WHERE telegram_id = ?", (user_id,))[0][
        0
    ]

    message = get_random_motivational_message(gender, called_from)
    if message:
        # username = user.username
        username = user_management.get_user_name(user_id)
        message = message.replace("(اسم المستخدم)", username)
        if isinstance(update, Update):
            await update.message.reply_text(message)
        else:
            await update.message.reply_text(message)
    context.user_data["button_clicks"] = 0


async def track_button_clicks(update: Update, context: CallbackContext, called_from):
    """Tracks button clicks and sends a motivational message when the threshold is reached."""
    user_data = context.user_data

    if "button_clicks" not in user_data:
        user_data["button_clicks"] = 0

    user_data["button_clicks"] += 1

    if user_data["button_clicks"] >= MESSAGES_TRIGGER_THRESHOLD:
        await send_motivational_message(update, context, called_from)
